import json
from datetime import datetime, timedelta
from io import BytesIO
from random import randint

import aiohttp_jinja2
from aiohttp import web
from aiohttp.web import Response
from openpyxl.workbook import Workbook
from sqlalchemy import or_

from entity.ServerEnum import ServerEnum
from lib.LogHelper import LogHelper
from config import INITIAL_SCORE, WIN_REWARD_SCORE, LOSE_PENALTY_SCORE
from convert.PlayerMatchData import TPlayerMatchData
from entity.PlayerRegInfo import PlayerInfo
from init_db import get_session
from lib.ScoreMaster import calculate_score
from lib.ServerManager import ServerManager
from lib.match_state import MatchConditionEx
from tables import *
from tables.WillMatch import DB_WillMatchs
from tables.PlayerNames import DB_PlayerNames

# 创建一个蓝图
bp = web.RouteTableDef()

# 定义处理程序
session = get_session()


def generate_numeric_code(length=6):
    return ''.join([str(randint(0, 9)) for _ in range(length)])


@bp.get('/')
async def index(request):
    try:
        context = {
            'title': 'Main Page',
            'content': 'Welcome to the main page!',
            'match_data': json.dumps([i.get_data() for i in session.query(DB_Matchs).all()])
        }
    except Exception as e:
        session.rollback()
        context = {
            'title': 'Main Page',
            'content': 'Welcome to the main page!',
            'match_data': json.dumps([i.get_data() for i in session.query(DB_Matchs).all()])
        }

    return aiohttp_jinja2.render_template('main.html', request, context)


@bp.get('/get_player_info/{player_id}')
async def get_player_info(request):
    player_id = request.match_info['player_id']
    session = get_session()
    query = session.query(DB_Player).filter(DB_Player.playerId == player_id)
    z1 = query.count()

    player_reg = PlayerInfo()
    player_reg.player_id = player_id

    if z1 == 0:  # 说明没注册
        query = session.query(DB_Verify).filter(DB_Verify.playerId == player_id)
        z = query.count()
        very = DB_Verify()
        if z == 0:
            very.code = generate_numeric_code()
            very.playerId = player_id
            session.add(very)
            session.commit()
        elif z == 1:
            very: DB_Verify = query.first()

        player_reg.verify_code = very.code
        player_reg.is_reg = False
    elif z1 > 0:
        player_reg.is_reg = True

    return Response(text=json.dumps(player_reg.__dict__))


@bp.get('/get_reg/{player_id}')
async def get_reg(req):
    player_id = req.match_info['player_id']
    with get_session() as session:
        query = session.query(DB_Verify).filter(DB_Verify.playerId == player_id)
        z = query.count()
        very = DB_Verify()
        if z == 0:
            very.code = generate_numeric_code()
            very.playerId = player_id
            session.add(very)
            session.commit()
        elif z == 1:
            very: DB_Verify = query.first()
        LogHelper.log(f"{player_id}, 验证码{very.code}")
    return Response(text=very.code)


@bp.get('/reg/{player_id}/{very_code}')
async def reg_player(request):
    # player_id = request.match_info['player_id']
    # try:
    #     sucess, reson = await add_player_code(player_id)
    # except Exception as e:
    #     sucess = False
    #     reson = str(e)
    # result_text = json.dumps({'success': sucess, 'reason': reson})
    return web.Response(text='not implement')


@bp.get('/GetVerifyCode')
async def get_verify_code(request):
    with get_session() as session:
        result = session.query(DB_Verify).filter(
            DB_Verify.until_time >= datetime.now()).all()
        list = {}
        for i in result:
            u: DB_Verify = i
            list[u.playerId] = u.code
        return web.Response(text=json.dumps(list))


@bp.get('/add_player_name/{player_id}/{name}')
async def add_player_name(request):
    player_id: str = request.match_info['player_id']
    name = request.match_info['name']
    LogHelper.log(f"添加 player_id 和 Name 指令：player_id:{player_id} name:{name}")

    if not player_id or not name:
        return Response(text='不得为空')

    session = get_session()
    z1 = session.query(DB_PlayerNames).filter(DB_PlayerNames.playerId == player_id).count()
    if z1 == 0:
        player_name = DB_PlayerNames()
        player_name.playerId = player_id
        player_name.PlayerName = name
        session.add(player_name)
        session.commit()
        LogHelper.log(f"已添加玩家名称:{name}")
    else:
        LogHelper.log("名字已存在")

    z2 = session.query(DB_Verify).filter(DB_Verify.playerId == player_id).count()
    if z2 == 0:
        verify = DB_Verify()
        verify.playerId = player_id
        verify.code = generate_numeric_code()
        session.add(verify)
        session.commit()

    return Response(text='success')


# @bp.get('/UploadMatchData/{json_str}')
# async def update_match_data(request):
#     print('update_match_data')
#     session.commit()
#
#     json_str = request.match_info['json_str']
#     data = json.loads(json_str)
#     k = json.loads(data)
#     players = k["_players"]
#     all_player_ids = list(players.keys())
#
#     # results = session.query(Player).filter(Player.playerId.in_(all_player_ids)).all()
#     for k, v in players.items():
#         q = session.query(Player).filter(Player.playerId == k)
#         if q.count() == 1:
#             player: Player = q.first()
#
#             player.team_damage += int(float(v["TKValue1"]))
#             player.damage += int(float(v["AttackValue1"]))
#             player.kill += int(v["KillNum1"])
#             player.death += int(v["DeadNum1"])
#
#     session.commit()


async def admin_add_player_name(player_id: str, name: str):
    session = get_session()
    LogHelper.log(f"player_id:{player_id} name{name}")
    z1 = session.query(DB_PlayerNames).filter(DB_PlayerNames.playerId == player_id).count()
    if z1 == 0:
        player_name = DB_PlayerNames()
        player_name.playerId = player_id
        player_name.PlayerName = name
        session.add(player_name)
        session.commit()
        LogHelper.log(f"已添加玩家名称:{name}")
    else:
        LogHelper.log("名字已存在")
    pass


"""
积分系统/计分系统
"""


@bp.post('/UploadMatchData')
async def update_match_data2(request):
    formatted_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[UploadMatchData|Right{formatted_time}]")

    data = await request.json()
    print(data)

    print('*' * 45)
    # for key, value in data.items():
    #     print(f"Key: {key}, Value: {value}")
    if not isinstance(data, dict):
        data = json.loads(data)
    match_data = DB_Matchs()
    match_data.left_scores = data["AttackScores"]
    match_data.right_scores = data["DefendScores"]
    match_data.left_players = list(set(data["AttackPlayerIds"]))
    match_data.right_players = list(set(data["DefendPlayerIds"]))
    match_data.left_win_rounds = data["AttackRound"]
    match_data.right_win_rounds = data["DefendRound"]
    match_data.server_name = data["ServerName"]
    match_data.tag = data["Tag"]
    playerData = data["_players"]
    match_data.player_data = playerData
    match_data.raw = data

    # #### 提前保存，防止数据异常 ####

    print('解析前的tag', match_data.tag, type(match_data.tag))
    if isinstance(match_data.tag, dict):
        match_data.tag = json.dumps(match_data.tag)
        print('属性判定', isinstance(match_data.tag, dict))
    session = get_session()
    session.add(match_data)
    session.commit()

    match_data.tag = json.loads(match_data.tag)
    if isinstance(match_data.tag, dict):
        is_match_ending = match_data.tag["IsMatchEnding"]
        round_count = int(match_data.tag["RoundCount"])
    else:
        round_count = 0
        is_match_ending = False
    if round_count < 3:
        LogHelper.log("未满3局")
        return
    if isinstance(match_data.tag, dict):
        match_data.tag = json.dumps(match_data.tag)
    # session.add(t)
    player_ids = list(playerData.keys())
    print(type(player_ids))
    print(player_ids)

    """
    保存数据
    """

    playerData_2: dict = playerData
    print('*' * 65)

    show_data = []
    team_a_data = []
    team_b_data = []
    player_score_dict = {}

    # 搜索玩家数据
    result = session.query(DB_PlayerData).filter(DB_PlayerData.playerId.in_(player_ids))

    result_player_ids = []
    for i in result:
        oldData: DB_PlayerData = i
        result_player_ids.append(oldData.playerId)
    #     找到之前没有数据的玩家
    missing_data = [value for key, value in playerData.items() if key not in result_player_ids]

    match_total_sum = match_data.get_total_data
    for i in result:
        oldData: DB_PlayerData = i
        k = TPlayerMatchData(playerData[oldData.playerId])
        k.set_total_matches(oldData.match)
        ## 计分系统
        # print(f"{k.player_id}: {k.win_rounds}")
        if not k.is_spectator_by_score():
            k.set_old_score(oldData.rank)

            if match_total_sum.attacker_rounds < match_total_sum.defender_rounds:
                if k.player_id in match_data.left_players:
                    print(
                        f"{k.player_name} .分数{match_total_sum.attacker_rounds}-{match_total_sum.defender_rounds}，false")
                    k.set_is_lose(True)
                elif k.player_id in match_data.right_players:
                    print(
                        f"{k.player_name} .分数{match_total_sum.attacker_rounds}-{match_total_sum.defender_rounds}，win")
                    k.set_is_lose(False)
            elif match_total_sum.attacker_rounds == match_total_sum.defender_rounds:
                print(f"{k.player_name} .分数{match_total_sum.attacker_rounds}-{match_total_sum.defender_rounds}")
                if k.player_id in match_data.left_players or k.player_id in match_data.right_players:
                    k.set_is_lose(False)
            elif match_total_sum.attacker_rounds > match_total_sum.defender_rounds:
                if k.player_id in match_data.left_players:
                    print(
                        f"{k.player_name} .分数{match_total_sum.attacker_rounds}-{match_total_sum.defender_rounds}，win")
                    k.set_is_lose(False)
                elif k.player_id in match_data.right_players:
                    print(
                        f"{k.player_name} .分数{match_total_sum.attacker_rounds}-{match_total_sum.defender_rounds}，false")
                    k.set_is_lose(True)
            # if k.win_rounds >= 3 or k.win >= 3:
            #     k.set_is_lose(False)
            # else:
            #     k.set_is_lose(True)
            sum_score, score_dict = calculate_score(match_total_sum, k)
            oldData.rank += sum_score
            player_score_dict[k.player_id] = score_dict

            k.set_new_score(oldData.rank)
            show_data.append(k)

            if k.player_id in match_data.left_players:
                team_a_data.append(k)
            else:
                team_b_data.append(k)
        else:
            LogHelper.log("跳过")

        oldData.add_match_data(k)
        # print(f"{oldData.rank}")
        oldData.playerName = k.player_name

    for i in missing_data:
        # print('test123')
        k = TPlayerMatchData(i)
        # print(k.__dict__)
        newData = DB_PlayerData()
        newData.clear_data()
        newData.playerId = k.player_id
        newData.playerName = k.player_name
        # 积分
        k.set_old_score(newData.rank)
        k.set_total_matches(1)
        if not k.is_spectator_by_score():
            if k.player_id in match_data.left_players:
                if match_total_sum.attacker_rounds < match_total_sum.defender_rounds:
                    k.set_is_lose(True)
                else:
                    k.set_is_lose(False)
            elif k.player_id in match_data.right_players:
                if match_total_sum.attacker_rounds > match_total_sum.defender_rounds:
                    k.set_is_lose(False)
                else:
                    k.set_is_lose(True)

            sum_score, score_dict = calculate_score(match_total_sum, k)
            newData.rank += sum_score

            player_score_dict[k.player_id] = score_dict

            k.set_new_score(newData.rank)
            show_data.append(k)

            if k.player_id in match_data.left_players:
                team_a_data.append(k)
            elif k.player_id in match_data.left_players:
                team_b_data.append(k)

        # 积分
        newData.add_match_data(k)

        session.add(newData)
    pass

    match_data.player_scores = json.dumps(player_score_dict)
    session.commit()

    # ################################### 设置比赛结束
    server_name = ServerManager.getServerName(ServerEnum.Server_1)
    server_name_withou_clear: str = match_data.server_name
    if '-' in server_name_withou_clear:
        server_name_withou_clear = server_name_withou_clear.split('-')[0]
    z = session.query(DB_WillMatchs).filter(DB_WillMatchs.server_name == server_name_withou_clear,
                                            DB_WillMatchs.match_id_2 == match_data.get_match_id)
    if z.count() > 0:
        for i in z:
            db_x: DB_WillMatchs = i
            db_x.is_finished = 1

            session.merge(db_x)
        session.commit()

    # (session.query(DB_WillMatchs).filter(DB_WillMatchs.match_id_2 == )
    # 保存数据到静态 , 转为有比赛结果
    MatchConditionEx.server_name = match_data.server_name
    MatchConditionEx.round_count = round_count
    MatchConditionEx.data = show_data
    MatchConditionEx.data2 = team_a_data + team_b_data
    MatchConditionEx.attacker_score = match_data.left_win_rounds
    MatchConditionEx.defender_score = match_data.right_win_rounds
    if is_match_ending:
        # 存放到 静态类中，让机器人输出
        MatchConditionEx.end_game = True
    LogHelper.log("数据已保存")

    return web.Response(text='123')


@bp.get('/Show_Last_Match')
async def Show_Last_Match(request):
    pass


@bp.get('/GetData')
async def GetData(request):
    # 创建一个Excel文件
    wb = Workbook()
    ws = wb.active

    one_hour_ago = datetime.now() - timedelta(hours=1)

    session.commit()
    latest_record = session.query(DB_Matchs).filter(
        DB_Matchs.time_match >= one_hour_ago,
    ).first()

    latest_match_data: DB_Matchs = latest_record
    print(latest_match_data)
    if latest_match_data is None:
        return web.Response(text='Data out dated')

    p = latest_match_data.raw
    # 找到 一小时内，含有相同玩家的所有记录
    all_player_ids = list(p.keys())
    some_records = session.query(DB_Matchs).filter(
        DB_Matchs.time_match >= one_hour_ago,
        or_(*[DB_Matchs.raw.like(f'%{s}%') for s in all_player_ids])
    )

    player_record = {}
    for record in some_records:
        match_data: DB_Matchs = record
        t = match_data.raw
        for k, v in t.items():
            u = TPlayerMatchData(v)
            if k not in player_record:
                player_record[k] = u
            else:
                old_data: TPlayerMatchData = player_record[k]
                if old_data.player_name is None:
                    old_data.player_name = u.player_name
                old_data.kill += u.kill
                old_data.death += u.death
                old_data.assist += u.assist
                old_data.win_rounds += u.win_rounds
                old_data.lose_rounds += u.lose_rounds
                old_data.damage += u.damage
                old_data.team_damage += u.team_damage
                old_data.horse_damage += u.horse_damage
                old_data.horse_tk += u.horse_damage

                player_record[k] = old_data
        pass

    ws.append(DB_PlayerData.GetHeaders)
    for key, record in player_record.items():
        record2: TPlayerMatchData = record
        ws.append(record2.get_data_list)

    # 美化
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # 将Excel文件内容写入BytesIO对象
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # 设置响应头信息
    response = web.StreamResponse()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename="{datetime.now().strftime("%m-%d-%H-%M-%S")}BTL.xlsx"'

    # 将BytesIO对象中的内容作为响应返回
    await response.prepare(request)
    await response.write(excel_data.read())

    return response


@bp.post('/GetDataEx')
async def GetDataEx(request):
    data = await request.json()
    print(data)
    # 创建一个Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = '玩家数据'
    session.commit()
    some_records = session.query(DB_Matchs).filter(
        DB_Matchs.id.in_(data)
    ).all()

    player_record = {}
    for record in some_records:
        match_data: DB_Matchs = record
        t = match_data.raw
        for k, v in t.items():
            u = TPlayerMatchData(v)
            if k not in player_record:
                player_record[k] = u
            else:
                old_data: TPlayerMatchData = player_record[k]
                if old_data.player_name is None:
                    old_data.player_name = u.player_name
                old_data.kill += u.kill
                old_data.death += u.death
                old_data.assist += u.assist
                old_data.win_rounds += u.win_rounds
                old_data.lose_rounds += u.lose_rounds
                old_data.damage += u.damage
                old_data.team_damage += u.team_damage
                old_data.horse_damage += u.horse_damage
                old_data.horse_tk += u.horse_tk

                player_record[k] = old_data
        pass

    ws.append(DB_PlayerData().GetHeaders())
    for key, record in player_record.items():
        record2: TPlayerMatchData = record
        ws.append(record2.get_data_list)

    # 美化
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # 附录信息
    new_sheet = wb.create_sheet(title='附录信息')
    new_sheet.append(['开始时间', ""])
    new_sheet.append(['结束时间', ""])

    # 将Excel文件内容写入BytesIO对象
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # 设置响应头信息
    response = web.StreamResponse()
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename="{datetime.now().strftime("%m-%d-%H-%M-%S")}BTL.xlsx"'

    # 将BytesIO对象中的内容作为响应返回
    await response.prepare(request)
    await response.write(excel_data.read())

    return response
    pass


# @bp.get('/v/{user_id}/{code}')
# async def validate_player(request):
#     user_id = request.match_info['user_id']
#     code = request.match_info['code']
if __name__ == '__main__':
    t = '''{
        "_players": {
            "2.0.0.76561199044372880": {
                "player_id": "2.0.0.76561199044372880",
                "AttackValue1": 0.0,
                "TKValue1": 0.0,
                "TKTimes1": 0.0,
                "TKPlayerNumbers1": 0.0,
                "AttackHorse1": 0,
                "TKHorse1": 0,
                "KillNum1": 0,
                "DeadNum1": 0
            },
            "2.0.0.76561198157649182": {
                "player_id": "2.0.0.76561198157649182",
                "AttackValue1": 100.0,
                "TKValue1": 0.0,
                "TKTimes1": 0.0,
                "TKPlayerNumbers1": 0.0,
                "AttackHorse1": 0,
                "TKHorse1": 0,
                "KillNum1": 0,
                "DeadNum1": 0
            }
        }
    }
    '''
    # k = json.loads(t)
    # players = k["_players"]
    # for k, v in players.items():
    #     print(v)
    #
    #    if round_count == 1:
    #         playerData_2: dict = playerData
    #         print('数据只有一项', '*' * 65)
    #
    #         show_data = []
    #         for j in playerData_2.values():
    #             k = TPlayerMatchData(j)
    #             show_data.append(k)
    #             pass
    #         # 保存数据到静态 , 转为有比赛结果
    #         MatchConditionEx.round_count = round_count
    #         MatchConditionEx.data = show_data
    #         LogHelper.log("数据已保存")
    #     else:
    #         # 加起来这些数据
    #         # 根据 RoundCount 取 n 个比赛数据
    #         print('round_count', round_count)
    #         result = (session.query(DB_Matchs).filter(DB_Matchs.server_name == t.server_name)
    #                   .order_by(desc(DB_Matchs.time_match))
    #                   .limit(round_count).all())
    #         show_data = []
    #         # 初始化数据
    #         first_match_obj: DB_Matchs = result[0]
    #         player_data_ex_cx: dict = first_match_obj.raw
    #         player_data_ex: dict = {}
    #         for k1, v1 in player_data_ex_cx.items():
    #             player_data_ex[k1] = TPlayerMatchData(v1)
    #
    #         for i in result[1:]:
    #             match_obj: DB_Matchs = i
    #             # match_obj.raw 直接就是dict
    #             playerData_2: dict = match_obj.raw
    #             print(playerData_2)
    #             print('+' * 65)
    #             for j in playerData_2.values():
    #                 k = TPlayerMatchData(j)
    #                 if k.player_id in player_data_ex:
    #                     cur_player_data = player_data_ex[k.player_id]
    #                     player_data_ex[k.player_id] = cur_player_data + k
    #                 else:
    #                     player_data_ex[k.player_id] = k
    #                 # 名字补全
    #                 await admin_add_player_name(k.player_id, k.player_name)

    #         pass
